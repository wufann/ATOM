#!/usr/bin/env python3
"""Build the OOB dashboard workbook and optionally upload it.

This uses the same raw OOT benchmark JSON inputs as
`.github/scripts/oot_benchmark_to_dashboard.py`, but writes them in the
Excel "Benchmark Summary" format expected by the OOB dashboard uploader:

- workbook filename: `benchmark_summary_<YYYYMMDD>.xlsx`
- one sheet per model
- columns: TP, ISL/OSL, Concurrency, TTFT, TPOT, E2EL, Output TPUT,
  Total TPUT, Per-GPU TPUT
"""

from __future__ import annotations

import argparse
import getpass
import importlib
import json
import os
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

VARIANT_RE = re.compile(r"-(mtp\d*)-")
DATE_RE = re.compile(r"(\d{8})")
INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]\*:/\\\?]")
DEFAULT_URL = "https://dashboard.amddevcloud.com/oob/oot_data_upload"
SKIP_FILENAMES = {
    "regression_report.json",
    "oot_benchmark_summary.json",
    "benchmark-action-input.json",
}
HEADERS = [
    "TP",
    "ISL/OSL",
    "Concurrency",
    "TTFT",
    "TPOT",
    "E2EL",
    "Output TPUT",
    "Total TPUT",
    "Per-GPU TPUT",
]


@dataclass(slots=True)
class BenchmarkRow:
    model: str
    tensor_parallel_size: int
    random_input_len: int
    random_output_len: int
    max_concurrency: int
    mean_ttft_ms: float | None
    mean_tpot_ms: float | None
    mean_e2el_ms: float | None
    output_throughput: float | None
    total_token_throughput: float | None
    total_token_throughput_per_gpu: float | None
    result_date: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert OOT benchmark JSON files to the OOB dashboard Excel format "
            "and optionally upload the workbook."
        )
    )
    parser.add_argument(
        "result_dir", help="Directory containing OOT benchmark JSON files."
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output workbook path. Defaults to "
            "`<result_dir>/benchmark_summary_<YYYYMMDD>.xlsx`."
        ),
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Report date to embed in the workbook filename (YYYYMMDD).",
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload the generated workbook to the OOB dashboard after building it.",
    )
    parser.add_argument("--url", default=DEFAULT_URL, help="Dashboard URL.")
    parser.add_argument(
        "--password",
        default=None,
        help="Dashboard password. Prefer --password-env or an interactive prompt.",
    )
    parser.add_argument(
        "--password-env",
        default="OOB_BENCHMARK_DATA_UPLOAD_PW",
        help="Environment variable used when --password is not provided.",
    )
    parser.add_argument(
        "--upload-button-text",
        default=None,
        help=(
            "Optional intermediate upload/submit button label to click after the file "
            "is selected and before the final database insert step."
        ),
    )
    parser.add_argument(
        "--finalize-button-text",
        default=r"insert\s+(?:into|to)\s+database",
        help=(
            "Regex for the final button used to commit parsed rows to the "
            "database. Pass an empty string to skip the final click."
        ),
    )
    parser.add_argument(
        "--done-text",
        default=None,
        help="Optional success text to wait for after the upload starts.",
    )
    parser.add_argument(
        "--timeout-ms",
        type=int,
        default=20000,
        help="Timeout for page operations in milliseconds.",
    )
    parser.add_argument(
        "--post-upload-wait-ms",
        type=int,
        default=5000,
        help="How long to wait after upload starts before taking the final screenshot.",
    )
    parser.add_argument(
        "--screenshot",
        default="oob_dashboard_upload.png",
        help="Path to save the final screenshot.",
    )
    parser.add_argument(
        "--show",
        action="store_true",
        help="Run headed so you can watch the browser actions.",
    )
    return parser.parse_args()


def derive_model_name(result_path: Path, payload: dict[str, Any]) -> str:
    display_name = payload.get("benchmark_model_name")
    if display_name:
        return str(display_name)

    model = str(payload.get("model_id", "")).split("/")[-1]
    if not model:
        model = result_path.stem

    match = VARIANT_RE.search(result_path.stem)
    if match:
        model = f"{model}-{match.group(1)}"

    return model


def is_publish_allowed(flag: object) -> bool:
    if flag is None:
        return True
    if isinstance(flag, bool):
        return flag
    return str(flag).strip().lower() not in {"0", "false", "no"}


def is_oob_dashboard_publish_allowed(payload: dict[str, Any]) -> bool:
    publish_flag = payload.get("oob_dashboard_publish_allowed")
    if publish_flag is None:
        publish_flag = payload.get("dashboard_publish_allowed")
    if publish_flag is None:
        return True
    return is_publish_allowed(publish_flag)


def parse_int(value: object, default: int = 0) -> int:
    try:
        return int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return default


def parse_float(value: object) -> float | None:
    try:
        if value in (None, ""):
            return None
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def extract_result_date(payload: dict[str, Any]) -> str | None:
    match = DATE_RE.search(str(payload.get("date", "")))
    return match.group(1) if match else None


def build_rows(result_dir: Path) -> list[BenchmarkRow]:
    rows: list[BenchmarkRow] = []
    for result_path in sorted(result_dir.glob("*.json")):
        if result_path.name in SKIP_FILENAMES:
            continue

        try:
            payload = json.loads(result_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            continue

        if not is_oob_dashboard_publish_allowed(payload):
            continue
        if "output_throughput" not in payload:
            continue

        tp = max(parse_int(payload.get("tensor_parallel_size"), 1), 1)
        total_tput = parse_float(payload.get("total_token_throughput"))
        total_tput_per_gpu = total_tput / tp if total_tput is not None else None

        rows.append(
            BenchmarkRow(
                model=derive_model_name(result_path, payload),
                tensor_parallel_size=tp,
                random_input_len=parse_int(payload.get("random_input_len")),
                random_output_len=parse_int(payload.get("random_output_len")),
                max_concurrency=parse_int(payload.get("max_concurrency")),
                mean_ttft_ms=parse_float(payload.get("mean_ttft_ms")),
                mean_tpot_ms=parse_float(payload.get("mean_tpot_ms")),
                mean_e2el_ms=parse_float(payload.get("mean_e2el_ms")),
                output_throughput=parse_float(payload.get("output_throughput")),
                total_token_throughput=total_tput,
                total_token_throughput_per_gpu=total_tput_per_gpu,
                result_date=extract_result_date(payload),
            )
        )

    return rows


def resolve_report_date(rows: Iterable[BenchmarkRow], override: str | None) -> str:
    if override:
        if not DATE_RE.fullmatch(override):
            raise SystemExit("--date must be in YYYYMMDD format")
        return override

    row_dates = sorted({row.result_date for row in rows if row.result_date})
    if row_dates:
        return row_dates[-1]

    return datetime.now(timezone.utc).strftime("%Y%m%d")


def sanitize_sheet_title(title: str, used_titles: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS_RE.sub("_", title).strip()
    cleaned = cleaned or "Model"
    base = cleaned[:31]
    candidate = base
    suffix = 1
    while candidate in used_titles:
        suffix_text = f"_{suffix}"
        candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def autosize_columns(worksheet) -> None:
    widths = {
        "A": 8,
        "B": 14,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 14,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def populate_sheet(worksheet, rows: list[BenchmarkRow]) -> None:
    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(
            [
                row.tensor_parallel_size,
                f"{row.random_input_len}/{row.random_output_len}",
                row.max_concurrency,
                row.mean_ttft_ms,
                row.mean_tpot_ms,
                row.mean_e2el_ms,
                row.output_throughput,
                row.total_token_throughput,
                row.total_token_throughput_per_gpu,
            ]
        )

    for column in ("D", "E", "F", "G", "H", "I"):
        for cell in worksheet[column][1:]:
            cell.number_format = "0.00"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def build_workbook(rows: list[BenchmarkRow], output_path: Path) -> Path:
    if not rows:
        raise SystemExit(
            "No eligible benchmark JSON files were found. "
            "Expected successful OOT result files with output_throughput."
        )

    grouped_rows: dict[str, list[BenchmarkRow]] = defaultdict(list)
    for row in rows:
        grouped_rows[row.model].append(row)

    workbook = Workbook()
    default_sheet = workbook.active
    used_titles: set[str] = set()

    for idx, model in enumerate(sorted(grouped_rows)):
        worksheet = default_sheet if idx == 0 else workbook.create_sheet()
        worksheet.title = sanitize_sheet_title(model, used_titles)
        populate_sheet(
            worksheet,
            sorted(
                grouped_rows[model],
                key=lambda row: (
                    row.tensor_parallel_size,
                    row.random_input_len,
                    row.random_output_len,
                    row.max_concurrency,
                ),
            ),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def resolve_password(args: argparse.Namespace) -> str:
    if args.password:
        return args.password

    env_value = os.getenv(args.password_env)
    if env_value:
        return env_value

    return getpass.getpass("Dashboard password: ")


def load_playwright_sync_api():
    try:
        return importlib.import_module("playwright.sync_api")
    except ImportError as exc:
        raise SystemExit(
            "Playwright is required for --upload.\n"
            "Install it with:\n"
            "  pip install playwright\n"
            "  playwright install chromium"
        ) from exc


def first_matching(
    candidates: Iterable[Any],
    *,
    require_visible: bool,
    timeout_error_type: type[Exception],
):
    for locator in candidates:
        try:
            if locator.count() == 0:
                continue
            candidate = locator.first
            if require_visible:
                candidate.wait_for(state="visible", timeout=1000)
            return candidate
        except timeout_error_type:
            continue
    return None


def upload_workbook(
    upload_file: Path,
    *,
    args: argparse.Namespace,
    password: str,
) -> None:
    playwright_sync = load_playwright_sync_api()
    sync_playwright = playwright_sync.sync_playwright
    timeout_error_type = playwright_sync.TimeoutError
    screenshot_path = Path(args.screenshot).expanduser().resolve()

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=not args.show)
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page()
        page.set_default_timeout(args.timeout_ms)

        def find_password_input():
            return first_matching(
                [
                    page.get_by_label(re.compile(r"(pw|password)", re.IGNORECASE)),
                    page.locator('input[type="password"]'),
                    page.locator('input[aria-label*="PW" i]'),
                    page.locator('[data-testid="stTextInput"] input'),
                ],
                require_visible=True,
                timeout_error_type=timeout_error_type,
            )

        def find_file_input():
            return first_matching(
                [
                    page.locator('input[type="file"]'),
                    page.locator('[data-testid="stFileUploader"] input[type="file"]'),
                ],
                require_visible=False,
                timeout_error_type=timeout_error_type,
            )

        def wait_for_file_input():
            deadline = time.monotonic() + args.timeout_ms / 1000
            while time.monotonic() < deadline:
                locator = find_file_input()
                if locator is not None:
                    return locator
                page.wait_for_timeout(500)
            return None

        def collect_button_texts() -> list[str]:
            texts: list[str] = []
            buttons = page.locator("button")
            try:
                count = min(buttons.count(), 20)
            except Exception:
                return texts

            for idx in range(count):
                try:
                    text = buttons.nth(idx).inner_text().strip()
                except Exception:
                    continue
                if text:
                    texts.append(text)
            return texts

        def find_action_button(patterns: list[str]):
            candidates = []
            for pattern in patterns:
                candidates.append(
                    page.get_by_role("button", name=re.compile(pattern, re.IGNORECASE))
                )
                candidates.append(page.locator(f"text=/{pattern}/i"))
            return first_matching(
                candidates,
                require_visible=True,
                timeout_error_type=timeout_error_type,
            )

        def wait_for_action_button(patterns: list[str], timeout_ms: int):
            deadline = time.monotonic() + timeout_ms / 1000
            while time.monotonic() < deadline:
                button = find_action_button(patterns)
                if button is not None:
                    return button
                page.wait_for_timeout(500)
            return None

        print(f"Opening {args.url}")
        page.goto(args.url, wait_until="domcontentloaded")
        page.wait_for_timeout(1500)

        password_input = find_password_input()
        if password_input is None:
            page.screenshot(path=str(screenshot_path), full_page=True)
            raise SystemExit(
                f"Could not find the password field. Screenshot saved to {screenshot_path}"
            )

        print("Filling dashboard password")
        password_input.fill(password)
        password_input.press("Enter")
        page.wait_for_timeout(1500)

        file_input = wait_for_file_input()
        if file_input is None:
            page.screenshot(path=str(screenshot_path), full_page=True)
            raise SystemExit(
                "Password was entered, but no file uploader appeared.\n"
                f"Visible buttons: {collect_button_texts()}\n"
                f"Screenshot saved to {screenshot_path}"
            )

        print(f"Selecting upload file: {upload_file}")
        file_input.set_input_files(str(upload_file))
        page.wait_for_timeout(1000)

        upload_patterns = (
            [args.upload_button_text]
            if args.upload_button_text
            else [
                r"upload",
                r"submit",
            ]
        )
        upload_button = wait_for_action_button(
            upload_patterns,
            timeout_ms=min(args.timeout_ms, 3000),
        )
        if upload_button is not None:
            label = upload_button.inner_text().strip() or "upload"
            print(f"Clicking upload button: {label}")
            upload_button.scroll_into_view_if_needed()
            upload_button.click()
        else:
            print(
                "No intermediate upload button found; continuing to the final insert step."
            )

        if args.finalize_button_text:
            finalize_button = wait_for_action_button(
                [args.finalize_button_text],
                timeout_ms=args.timeout_ms,
            )
            if finalize_button is None:
                page.screenshot(path=str(screenshot_path), full_page=True)
                raise SystemExit(
                    "Excel file was selected, but the final insert button did not appear.\n"
                    f"Visible buttons: {collect_button_texts()}\n"
                    f"Screenshot saved to {screenshot_path}"
                )

            label = finalize_button.inner_text().strip() or args.finalize_button_text
            print(f"Clicking finalize button: {label}")
            finalize_button.scroll_into_view_if_needed()
            finalize_button.click()
            page.wait_for_timeout(1000)

        if args.done_text:
            print(f"Waiting for completion text: {args.done_text}")
            page.get_by_text(re.compile(args.done_text, re.IGNORECASE)).first.wait_for(
                state="visible",
                timeout=args.timeout_ms,
            )
        else:
            page.wait_for_timeout(args.post_upload_wait_ms)

        page.screenshot(path=str(screenshot_path), full_page=True)
        print(f"Upload flow completed. Screenshot saved to {screenshot_path}")

        context.close()
        browser.close()


def main() -> int:
    args = parse_args()
    result_dir = Path(args.result_dir).expanduser().resolve()
    if not result_dir.is_dir():
        print(f"Result directory does not exist: {result_dir}", file=sys.stderr)
        return 2

    rows = build_rows(result_dir)
    report_date = resolve_report_date(rows, args.date)
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else result_dir / f"benchmark_summary_{report_date}.xlsx"
    )
    workbook_path = build_workbook(rows, output_path)

    model_count = len({row.model for row in rows})
    print(
        f"Generated workbook with {len(rows)} benchmark rows across {model_count} model sheets: "
        f"{workbook_path}"
    )

    if args.upload:
        password = resolve_password(args)
        upload_workbook(workbook_path, args=args, password=password)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
