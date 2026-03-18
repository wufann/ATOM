#!/usr/bin/env python3
"""
Parse PyTorch profiler trace JSON to extract kernel information.

Usage:
    python parse_trace.py <trace.json> [--layer N]
"""

import json
import gzip
import sys
import bisect
import argparse
import re
import os
from glob import glob
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import Workbook

# Modules to filter out (no corresponding GPU kernel in decode)
FILTER_OUT = ["fill_"]

# Sampling-related modules and low-level ops to filter out in prefill
FILTER_OUT_PREFILL = ["aiter::mixed_sample"]


# =============================================================================
# Utility Functions
# =============================================================================


def load_trace(filepath: str) -> Dict[str, Any]:
    """Load trace JSON file (supports .gz)."""
    opener = gzip.open if filepath.endswith(".gz") else open
    with opener(filepath, "rt", encoding="utf-8") as f:
        return json.load(f)


def is_kernel_launch(name: str) -> bool:
    """Check if name is a kernel launch (contains 'launch' and 'kernel')."""
    n = name.lower()
    return "launch" in n and "kernel" in n


def should_filter(name: str) -> bool:
    """Check if module should be filtered out."""
    return any(f in name for f in FILTER_OUT)


def should_filter_prefill(name: str) -> bool:
    """Check if module should be filtered out in prefill (sampling ops)."""
    return any(f in name for f in FILTER_OUT_PREFILL)


def is_strict_norm_name(name: str) -> bool:
    """Match norm module names strictly, not by substring."""
    if not isinstance(name, str):
        return False
    n = name.strip().lower()
    return n == "layernorm" or n == "rmsnorm"


def extract_model_name_from_trace_filename(filepath: str) -> Optional[str]:
    """
    Extract model name from trace filename prefix before `_ts_`.
    Examples:
      - Meta-Llama-3.1-8B-Instruct_ts_... -> Meta-Llama-3.1-8B-Instruct
      - capture_graph_Meta-Llama-3.1-8B-Instruct_ts_... -> Meta-Llama-3.1-8B-Instruct
    """
    base = os.path.basename(filepath)
    if "_ts_" not in base:
        return None
    prefix = base.split("_ts_", 1)[0]
    if prefix.startswith("capture_graph_"):
        prefix = prefix[len("capture_graph_") :]
    return prefix or None


def find_capture_graph_trace_path(run_trace_path: str) -> Optional[str]:
    """
    Find capture graph trace file in the same directory as the run trace.
    Pattern: capture_graph_<model>_ts_*.pt.trace.json[.gz]
    """
    model_name = extract_model_name_from_trace_filename(run_trace_path)
    if not model_name:
        return None
    trace_dir = os.path.dirname(run_trace_path) or "."
    pattern = os.path.join(trace_dir, f"capture_graph_{model_name}_ts_*.pt.trace.json*")
    candidates = sorted(glob(pattern), key=os.path.getmtime, reverse=True)
    if not candidates:
        return None
    run_abs = os.path.abspath(run_trace_path)
    for fp in candidates:
        if os.path.abspath(fp) != run_abs:
            return fp
    return None


_LLVM_CXXFILT_PATH: Optional[str] = None
_DEMANGLE_CACHE: Dict[str, str] = {}


def _find_llvm_cxxfilt() -> Optional[str]:
    """Find llvm-cxxfilt binary (supports HIP/ROCm mangled names)."""
    global _LLVM_CXXFILT_PATH
    if _LLVM_CXXFILT_PATH is not None:
        return _LLVM_CXXFILT_PATH or None

    import shutil
    import subprocess

    # Check PATH first
    path = shutil.which("llvm-cxxfilt")
    if path:
        _LLVM_CXXFILT_PATH = path
        return path

    # Check known install paths before resorting to find
    known_paths = [
        "/opt/rocm/llvm/bin/llvm-cxxfilt",
        "/usr/bin/llvm-cxxfilt",
        "/usr/local/bin/llvm-cxxfilt",
    ]
    for p in known_paths:
        if os.path.isfile(p):
            _LLVM_CXXFILT_PATH = p
            return p

    # Fallback: search common locations with depth limit
    search_dirs = ["/root/.triton/llvm", "/opt/rocm"]
    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        try:
            result = subprocess.run(
                [
                    "find",
                    d,
                    "-maxdepth",
                    "5",
                    "-name",
                    "llvm-cxxfilt",
                    "-type",
                    "f",
                    "-print",
                    "-quit",
                ],
                capture_output=True,
                text=True,
                timeout=5,
            )
            found = result.stdout.strip()
            if found:
                _LLVM_CXXFILT_PATH = found
                return found
        except (subprocess.TimeoutExpired, OSError):
            continue

    _LLVM_CXXFILT_PATH = ""
    return None


def _demangle_batch(names: list[str]) -> None:
    """Batch demangle C++ mangled names using a single llvm-cxxfilt call."""
    mangled = [n for n in names if n.startswith("_Z") and n not in _DEMANGLE_CACHE]
    if not mangled:
        return

    cxxfilt = _find_llvm_cxxfilt()
    if not cxxfilt:
        for n in mangled:
            _DEMANGLE_CACHE[n] = n
        return

    import subprocess

    try:
        result = subprocess.run(
            [cxxfilt],
            input="\n".join(mangled),
            capture_output=True,
            text=True,
            timeout=30,
        )
        demangled_list = result.stdout.strip().split("\n")
        for orig, dem in zip(mangled, demangled_list):
            _DEMANGLE_CACHE[orig] = dem.strip() if dem.strip() else orig
    except (subprocess.TimeoutExpired, OSError):
        for n in mangled:
            _DEMANGLE_CACHE[n] = n


def _demangle_kernel_name(name: str) -> str:
    """Demangle C++ mangled kernel name (uses cache populated by _demangle_batch)."""
    if name in _DEMANGLE_CACHE:
        return _DEMANGLE_CACHE[name]
    _DEMANGLE_CACHE[name] = name
    return name


def write_breakdown_xlsx(
    output_xlsx: str,
    rows: List[List[Any]],
    sheet_name: str,
    avg_rows: Optional[List[List[Any]]] = None,
) -> None:
    """
    Write XLSX breakdown with columns:
    cpu_module, gpu_kernel, duration_us, pct%, sum per module, module_pct%,
    avg duration_us, avg sum per module.

    Also writes a 'kernel_summary' sheet with aggregated kernel statistics.

    The 1st/5th columns are merged for contiguous identical modules.
    AVG columns are appended to the right in the same table.
    """
    wb = Workbook()
    # Batch demangle all kernel names upfront (single subprocess call)
    all_kernels = [r[1] for r in rows]
    if avg_rows:
        all_kernels.extend(r[1] for r in avg_rows)
    _demangle_batch(all_kernels)

    ws = wb.active
    ws.title = sheet_name
    ws.append(
        [
            "cpu_module",
            "gpu_kernel",
            "duration_us",
            "pct%",
            "sum per module",
            "module_pct%",
            "avg duration_us",
            "avg sum per module",
        ]
    )

    total_duration = sum(float(r[2]) for r in rows) if rows else 0.0

    def build_groups(block_rows: List[List[Any]]) -> List[Tuple[int, int, str, float]]:
        groups: List[Tuple[int, int, str, float]] = []
        i = 0
        while i < len(block_rows):
            mod = block_rows[i][0]
            j = i + 1
            total = float(block_rows[i][2])
            while j < len(block_rows) and block_rows[j][0] == mod:
                total += float(block_rows[j][2])
                j += 1
            groups.append((i, j - 1, mod, total))
            i = j
        return groups

    main_groups = build_groups(rows) if rows else []
    renamed_group_mods = [g[2] for g in main_groups]
    seen_rmsnorm = 0
    for gi, mod in enumerate(renamed_group_mods):
        if is_strict_norm_name(mod):
            if seen_rmsnorm == 0:
                renamed_group_mods[gi] = "input_layernorm"
            elif seen_rmsnorm == 1:
                renamed_group_mods[gi] = "post_attn_layernorm"
            seen_rmsnorm += 1

    avg_sum_by_row: Dict[int, float] = {}
    if avg_rows:
        avg_groups = build_groups(avg_rows)
        for start, end, _, total in avg_groups:
            for i in range(start, end + 1):
                avg_sum_by_row[i] = total

    data_start_row = ws.max_row + 1
    for gi, (start, end, _, mod_total) in enumerate(main_groups):
        renamed_mod = renamed_group_mods[gi]
        mod_pct = (mod_total / total_duration * 100) if total_duration > 0 else 0
        for idx in range(start, end + 1):
            _, kernel, dur = rows[idx]
            dur_f = float(dur)
            pct = (dur_f / total_duration * 100) if total_duration > 0 else 0
            avg_dur = (
                float(avg_rows[idx][2]) if avg_rows and idx < len(avg_rows) else ""
            )
            avg_sum = avg_sum_by_row.get(idx, "")
            ws.append(
                [
                    renamed_mod,
                    _demangle_kernel_name(kernel),
                    dur,
                    round(pct, 1),
                    mod_total,
                    round(mod_pct, 1),
                    avg_dur,
                    avg_sum,
                ]
            )

    for start, end, _, _ in main_groups:
        if end > start:
            r1 = data_start_row + start
            r2 = data_start_row + end
            ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
            ws.merge_cells(start_row=r1, start_column=5, end_row=r2, end_column=5)
            ws.merge_cells(start_row=r1, start_column=6, end_row=r2, end_column=6)
            if avg_rows:
                ws.merge_cells(start_row=r1, start_column=8, end_row=r2, end_column=8)

    total_avg_duration = sum(float(r[2]) for r in avg_rows) if avg_rows else ""
    ws.append(["TOTAL", "", total_duration, 100.0, "", 100.0, total_avg_duration, ""])

    # --- Kernel Summary Sheet ---
    if rows:
        _write_kernel_summary_sheet(wb, rows, total_duration)

    wb.save(output_xlsx)


def _write_kernel_summary_sheet(
    wb: "Workbook",
    rows: List[List[Any]],
    total_duration: float,
) -> None:
    """Write a kernel_summary sheet: aggregated stats by kernel name."""
    ws = wb.create_sheet("kernel_summary")
    ws.append(["gpu_kernel", "calls", "total_duration_us", "avg_duration_us", "pct%"])

    kernel_stats: Dict[str, List[float]] = {}
    for _, kernel, dur in rows:
        short_name = _demangle_kernel_name(kernel)
        kernel_stats.setdefault(short_name, []).append(float(dur))

    # Sort by total duration descending
    sorted_kernels = sorted(kernel_stats.items(), key=lambda x: sum(x[1]), reverse=True)

    for kernel_name, durations in sorted_kernels:
        total_dur = sum(durations)
        count = len(durations)
        avg_dur = total_dur / count
        pct = (total_dur / total_duration * 100) if total_duration > 0 else 0
        ws.append(
            [kernel_name, count, round(total_dur, 3), round(avg_dur, 3), round(pct, 1)]
        )


def _normalize_module_for_avg(name: str) -> str:
    if not isinstance(name, str):
        return str(name)
    return re.sub(r"model\.layers\.\d+\.", "model.layers.*.", name)


def build_avg_rows_from_layers(
    layer_rows_list: List[Tuple[int, List[List[Any]]]],
    section_name: str,
) -> Optional[List[List[Any]]]:
    """
    Build AVG rows across layers with fallback:
    1) try contiguous layers from avg_start_layer.
    2) if mismatch, retry every other layer: start, start+2, start+4, ...
    Returns None if still not alignable.
    """
    if not layer_rows_list:
        return []

    def _try_build(
        selected_layers: List[Tuple[int, List[List[Any]]]],
    ) -> Tuple[Optional[List[List[Any]]], Optional[int]]:
        base_layer, base_rows = selected_layers[0]
        base_sig = [(_normalize_module_for_avg(r[0]), r[1]) for r in base_rows]

        for layer_idx, rows in selected_layers[1:]:
            sig = [(_normalize_module_for_avg(r[0]), r[1]) for r in rows]
            if sig != base_sig:
                return None, layer_idx

        n = len(selected_layers)
        avg_rows: List[List[Any]] = []
        for i, (_, kernel) in enumerate(base_sig):
            # Keep display style from base layer rows.
            display_mod = base_rows[i][0]
            avg_dur = (
                sum(float(selected_layers[layer_i][1][i][2]) for layer_i in range(n))
                / n
            )
            avg_rows.append([display_mod, kernel, avg_dur])
        return avg_rows, None

    start_layer = layer_rows_list[0][0]
    avg_rows, bad_layer = _try_build(layer_rows_list)
    if avg_rows is not None:
        return avg_rows

    print(
        f"{section_name} avg mismatch: layer {bad_layer} does not match layer {start_layer} layout."
    )
    fallback_layers = [
        item for item in layer_rows_list if (item[0] - start_layer) % 2 == 0
    ]
    fallback_indices = [str(layer_idx) for layer_idx, _ in fallback_layers]
    print(
        f"{section_name} avg retry with every other layer: {'.'.join(fallback_indices)}"
    )
    if len(fallback_layers) < 2:
        print(f"{section_name} avg skipped: fallback has fewer than 2 layers.")
        return None

    avg_rows, bad_layer = _try_build(fallback_layers)
    if avg_rows is not None:
        return avg_rows

    print(f"{section_name} avg skipped: still mismatch at layer {bad_layer}.")
    return None


# =============================================================================
# Optimized Event Index for fast time-range queries
# =============================================================================


class EventIndex:
    """Pre-indexed events for fast time-range queries."""

    def __init__(self, events: List[Dict]):
        # Filter duration events only
        self.duration_events = [e for e in events if e.get("ph") == "X"]
        self.duration_events.sort(key=lambda x: x["ts"])
        self.ts_list = [e["ts"] for e in self.duration_events]

        # Pre-compute kernel launch flags and prefix sum
        self._is_kernel_launch = [
            is_kernel_launch(e.get("name", "")) for e in self.duration_events
        ]
        self._kernel_prefix_sum = [0]
        for is_kl in self._is_kernel_launch:
            self._kernel_prefix_sum.append(
                self._kernel_prefix_sum[-1] + (1 if is_kl else 0)
            )

    def events_in_range(self, start_ts: float, end_ts: float) -> List[Dict]:
        """Get all duration events within [start_ts, end_ts]."""
        left = bisect.bisect_left(self.ts_list, start_ts)
        right = bisect.bisect_right(self.ts_list, end_ts)
        return [
            e
            for e in self.duration_events[left:right]
            if e["ts"] + e.get("dur", 0) <= end_ts
        ]

    def count_kernel_launches_in_range(self, start_ts: float, end_ts: float) -> int:
        """Count kernel launches within time range (fast using prefix sum)."""
        left = bisect.bisect_left(self.ts_list, start_ts)
        right = bisect.bisect_right(self.ts_list, end_ts)
        count = 0
        for i in range(left, right):
            e = self.duration_events[i]
            if e["ts"] + e.get("dur", 0) <= end_ts and self._is_kernel_launch[i]:
                count += 1
        return count

    def get_direct_children(self, parent: Dict) -> List[Dict]:
        """Get direct children of parent event (optimized)."""
        p_ts = parent["ts"]
        p_end = p_ts + parent.get("dur", 0)

        # Get candidates in parent's time range
        candidates = [e for e in self.events_in_range(p_ts, p_end) if e is not parent]

        if not candidates:
            return []

        # Filter to direct children only (not nested in other candidates)
        # Sort by duration descending - larger events are potential parents
        candidates_sorted = sorted(candidates, key=lambda x: -x.get("dur", 0))

        direct = []
        for i, c in enumerate(candidates_sorted):
            c_ts, c_dur = c["ts"], c.get("dur", 0)
            c_end = c_ts + c_dur
            # Check if c is nested inside any larger candidate
            is_nested = False
            for j in range(i):  # Only check larger (earlier in sorted list)
                o = candidates_sorted[j]
                o_ts = o["ts"]
                o_end = o_ts + o.get("dur", 0)
                if c_ts >= o_ts and c_end <= o_end:
                    is_nested = True
                    break
            if not is_nested:
                direct.append(c)

        return sorted(direct, key=lambda x: x["ts"])

    def count_kernel_launches(self, event: Dict) -> int:
        """Count kernel launches within event's time range."""
        e_ts = event["ts"]
        e_end = e_ts + event.get("dur", 0)
        return self.count_kernel_launches_in_range(e_ts, e_end)

    def has_kernel_launch(self, event: Dict) -> bool:
        """Check if event contains any kernel launch."""
        return self.count_kernel_launches(event) > 0


# =============================================================================
# Parse Functions
# =============================================================================


def parse_prefill(events: List[Dict], output_xlsx: str, target_layer: int = 3) -> None:
    """
    Parse prefill phase from a run trace (no warmup mixed in this trace).
    """
    # CPU side prefill annotations.
    # Matches "prefill[bs=1 tok=115 ctx=115]" format.
    prefills = [
        e
        for e in events
        if e.get("name", "").startswith("prefill[")
        and e.get("ph") == "X"
        and e.get("cat") == "user_annotation"
    ]
    prefills = sorted(prefills, key=lambda x: x["ts"])

    if not prefills:
        print("No prefill (user_annotation) events found.")
        write_breakdown_xlsx(output_xlsx, [], sheet_name="prefill")
        return

    actual_prefill_idx = 0
    actual_prefill = prefills[actual_prefill_idx]
    print(f"Found {len(prefills)} prefill events (user_annotation)")
    print("Using first prefill event (run trace has no warmup phase).")
    print(
        f"Using prefill[{actual_prefill_idx}] "
        f"(ts={actual_prefill.get('ts', 0):.0f}, dur={actual_prefill.get('dur', 0):.0f})"
    )

    # Build prefill hierarchy on the same thread as the selected CPU prefill.
    # Using thread affinity is more robust than category-only filtering.
    prefill_tid = actual_prefill.get("tid")
    prefill_pid = actual_prefill.get("pid")
    prefill_hierarchy_events = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("tid") == prefill_tid
        and e.get("pid") == prefill_pid
    ]
    # Build index once for fast subtree queries in prefill parsing.
    prefill_idx = EventIndex(prefill_hierarchy_events)
    level1_children = prefill_idx.get_direct_children(actual_prefill)
    print(
        f"Prefill level 1 (same thread pid={prefill_pid}, tid={prefill_tid}): "
        f"{len(level1_children)} nodes"
    )

    # Keep only level2 children that have kernel launch in their subtree.
    launch_level2_items = []
    for l1 in level1_children:
        l1_name = l1.get("name", "<unknown>")
        level2_children = prefill_idx.get_direct_children(l1)
        level2_with_launch = [
            l2 for l2 in level2_children if prefill_idx.has_kernel_launch(l2)
        ]
        for l2 in level2_with_launch:
            launch_level2_items.append(
                {
                    "level1_name": l1_name,
                    "level2_event": l2,
                }
            )

    print(f"Level2 children with kernel launch: {len(launch_level2_items)}")

    # Layer extraction by rmsnorm positions:
    # each layer has 2 rmsnorm modules, layer N starts at rmsnorm index 2*N (0-based).
    all_norm_indices = [
        i
        for i, item in enumerate(launch_level2_items)
        if is_strict_norm_name(item["level2_event"].get("name", ""))
    ]

    # Build launch->kernel mapping by correlation id.
    # Build launch candidates from current prefill thread/range once.
    runtime_launches = [
        e
        for e in prefill_hierarchy_events
        if e.get("cat") == "cuda_runtime" and is_kernel_launch(e.get("name", ""))
    ]
    runtime_launches.sort(key=lambda x: x.get("ts", 0))
    runtime_launch_ts = [e.get("ts", 0) for e in runtime_launches]

    item_corrs: List[List[Any]] = []
    corr_needed = set()
    for item in launch_level2_items:
        l2 = item["level2_event"]
        l2_start = l2.get("ts", 0)
        l2_end = l2_start + l2.get("dur", 0)

        left = bisect.bisect_left(runtime_launch_ts, l2_start)
        right = bisect.bisect_right(runtime_launch_ts, l2_end)
        launches_in_l2 = runtime_launches[left:right]
        curr_corrs = []
        for launch in launches_in_l2:
            corr = (launch.get("args") or {}).get("correlation")
            if corr is not None:
                corr_needed.add(corr)
                curr_corrs.append(corr)
        item_corrs.append(curr_corrs)

    # Build kernel index only for correlations we actually need.
    kernel_by_corr: Dict[Any, List[Dict]] = {}
    if corr_needed:
        for e in events:
            if e.get("ph") != "X" or e.get("cat") != "kernel":
                continue
            corr = (e.get("args") or {}).get("correlation")
            if corr is None or corr not in corr_needed:
                continue
            kernel_by_corr.setdefault(corr, []).append(e)
        for corr in kernel_by_corr:
            kernel_by_corr[corr].sort(key=lambda x: x.get("ts", 0))

    item_kernels: List[List[Dict[str, Any]]] = []
    for corrs in item_corrs:
        kernels = []
        for corr in corrs:
            for k in kernel_by_corr.get(corr, []):
                kernels.append({"name": k.get("name", "N/A"), "dur": k.get("dur", 0)})
        item_kernels.append(kernels)

    def _resolve_moe_child_name_prefill(event: Dict[str, Any]) -> str:
        mod_name = event.get("name", "<unknown>")
        if "moe" not in mod_name.lower():
            return mod_name
        children = prefill_idx.get_direct_children(event)
        children_with_launch = [c for c in children if prefill_idx.has_kernel_launch(c)]
        if children_with_launch:
            return children_with_launch[0].get("name", mod_name)
        return mod_name

    def build_rows_from_item_range(start: int, end: int) -> List[List[Any]]:
        rows = []
        for i in range(start, end):
            item = launch_level2_items[i]
            mod_name = _resolve_moe_child_name_prefill(item["level2_event"])
            if should_filter_prefill(mod_name):
                continue
            kernels = [k for k in item_kernels[i] if k["name"] not in ("", "N/A")]
            if not kernels:
                continue
            rows.extend(process_module(mod_name, len(kernels), 0, kernels))
        return rows

    _extract_layer_and_write(
        all_norm_indices,
        len(launch_level2_items),
        target_layer,
        "Prefill",
        "prefill",
        build_rows_from_item_range,
        output_xlsx,
    )


def clean_module_name(name: str, mapped_kernel_name: str = "") -> str:
    """Clean and simplify module name."""
    # Runtime launch wrappers should display the actual launched operator name.
    if "hipmodulelaunchkernel" in name.lower() and mapped_kernel_name not in (
        "",
        "N/A",
    ):
        name = _demangle_kernel_name(mapped_kernel_name)

    # Demangle mangled C++ names
    if name.startswith("_Z"):
        name = _demangle_kernel_name(name)

    # Remove 'aiter::' prefix if present
    if name.startswith("aiter::"):
        name = name[7:]  # len('aiter::') == 7

    return name


def _extract_layer_and_write(
    all_norm_indices: List[int],
    total_module_count: int,
    target_layer: int,
    section_name: str,
    sheet_name: str,
    build_rows_fn,
    output_xlsx: str,
) -> None:
    """
    Shared layer-extraction, AVG computation, and XLSX write logic
    used by both parse_prefill and parse_decode.

    Args:
        all_norm_indices: indices of all norm modules (including final layernorm)
        total_module_count: total number of modules (used as fallback final_norm_idx)
        target_layer: which layer to extract
        section_name: "Prefill" or "Decode" for print messages
        sheet_name: XLSX sheet name
        build_rows_fn: callable(start, end) -> List[List[Any]]
        output_xlsx: output file path
    """
    norm_indices = all_norm_indices[:-1] if len(all_norm_indices) > 0 else []
    print(
        f"Found {len(all_norm_indices)} norm modules "
        f"({len(norm_indices)} used for layer split, excluding final layernorm)"
    )

    TARGET_LAYER = target_layer
    norm_start_idx = TARGET_LAYER * 2
    norm_end_idx = (TARGET_LAYER + 1) * 2
    final_norm_idx = (
        all_norm_indices[-1] if len(all_norm_indices) > 0 else total_module_count
    )

    mod_start = 0
    mod_end = 0
    if norm_start_idx >= len(norm_indices):
        print(f"Not enough norms for layer {TARGET_LAYER}")
        if sheet_name == "prefill":
            print(
                f"Not enough rmsnorm modules for layer {TARGET_LAYER}, writing empty XLSX"
            )
        write_breakdown_xlsx(output_xlsx, [], sheet_name=sheet_name)
        return
    else:
        mod_start = norm_indices[norm_start_idx]
        mod_end = (
            norm_indices[norm_end_idx]
            if norm_end_idx < len(norm_indices)
            else final_norm_idx
        )
        print(
            f"Layer {TARGET_LAYER}: modules [{mod_start}:{mod_end}] "
            f"(norms at indices {norm_start_idx}, {norm_start_idx+1})"
        )

    avg_start_layer = TARGET_LAYER
    avg_end_layer = (len(norm_indices) - 1) // 2
    if avg_start_layer <= avg_end_layer:
        print(
            f"Target layer: {TARGET_LAYER}; AVG layers: [{avg_start_layer}..{avg_end_layer}]"
        )
    else:
        print(
            f"Target layer: {TARGET_LAYER}; AVG layers: disabled (no eligible layers)"
        )

    # Target layer rows.
    rows = build_rows_fn(mod_start, mod_end)
    print(f"Layer {TARGET_LAYER} rows: {len(rows)}")

    # AVG rows.
    avg_rows = None
    avg_layer_rows: List[Tuple[int, List[List[Any]]]] = []
    layer = avg_start_layer
    while 2 * layer < len(norm_indices):
        s = norm_indices[2 * layer]
        e_idx = 2 * (layer + 1)
        e = norm_indices[e_idx] if e_idx < len(norm_indices) else final_norm_idx
        avg_layer_rows.append((layer, build_rows_fn(s, e)))
        layer += 1
    if avg_layer_rows:
        avg_rows = build_avg_rows_from_layers(avg_layer_rows, section_name)
        if avg_rows is not None:
            print(f"{section_name} avg rows: {len(avg_rows)}")

    write_breakdown_xlsx(output_xlsx, rows, sheet_name=sheet_name, avg_rows=avg_rows)
    print(f"Layer {TARGET_LAYER} modules: {mod_end - mod_start}")
    print(f"XLSX written to: {output_xlsx} ({len(rows)} rows)")


def process_module(
    mod_name: str, kernel_count: int, start_gpu_idx: int, gpu_kernels: List[Dict]
) -> List[List]:
    """Process a module and return [display_name, gpu_kernel_name, gpu_dur] rows."""
    rows = []
    for i in range(kernel_count):
        gpu_idx = start_gpu_idx + i
        gpu_kernel_name = "N/A"
        gpu_dur = 0
        if gpu_idx < len(gpu_kernels):
            gpu = gpu_kernels[gpu_idx]
            gpu_kernel_name = gpu.get("name", "N/A")
            gpu_dur = gpu.get("dur", 0)
        display_name = clean_module_name(mod_name, gpu_kernel_name)
        rows.append([display_name, gpu_kernel_name, gpu_dur])
    return rows


def parse_decode(
    run_events: List[Dict],
    capture_events: List[Dict],
    output_xlsx: str,
    target_layer: int = 3,
) -> None:
    """
    Parse decode phase:
    - use run trace for first decode and real kernel timings
    - use capture trace for capture_graph module hierarchy

    Output CSV columns: cpu_module, gpu_kernel, duration_us
    """
    print("Building event index...")

    # Find GPU-annotated decode events (cat='gpu_user_annotation')
    decodes = [
        e
        for e in run_events
        if e.get("name", "").startswith("decode[")
        and e.get("ph") == "X"
        and e.get("cat") == "gpu_user_annotation"
    ]
    decodes = sorted(decodes, key=lambda x: x["ts"])

    if not decodes:
        print("No decode (gpu_user_annotation) events found.")
        return

    first_ds = decodes[0]
    first_ds_name = first_ds.get("name", "")
    target_bs: Optional[int] = None
    bs_match = re.search(r"bs=(\d+)", first_ds_name)
    if bs_match:
        target_bs = int(bs_match.group(1))
        target_cg_name = f"capture_graph_bs_{target_bs}"
    else:
        target_cg_name = "capture_graph"

    print(f"First decode: {first_ds_name}")
    print(f"Looking for: {target_cg_name}")

    # Find matching capture_graph
    capture_graphs = [
        e
        for e in capture_events
        if e.get("name") == target_cg_name and e.get("ph") == "X"
    ]
    if not capture_graphs and target_bs is not None:
        # Prefer the largest capture_graph_bs_K where K < target_bs.
        lower_bs_candidates: List[Tuple[int, Dict[str, Any]]] = []
        for e in capture_events:
            if e.get("ph") != "X":
                continue
            n = e.get("name", "")
            m = re.match(r"^capture_graph_bs_(\d+)$", n)
            if not m:
                continue
            k = int(m.group(1))
            if k < target_bs:
                lower_bs_candidates.append((k, e))
        if lower_bs_candidates:
            best_bs = max(k for k, _ in lower_bs_candidates)
            capture_graphs = sorted(
                [e for k, e in lower_bs_candidates if k == best_bs],
                key=lambda x: x.get("ts", 0),
            )
            print(f"No exact match, using nearest lower capture_graph_bs_{best_bs}")
    if not capture_graphs:
        # Fallback: find any capture_graph
        capture_graphs = [
            e
            for e in capture_events
            if e.get("name", "").startswith("capture_graph") and e.get("ph") == "X"
        ]
        capture_graphs = sorted(capture_graphs, key=lambda x: x["ts"])
        print("No exact match, using first capture_graph")

    if not capture_graphs:
        print("No capture_graph events found.")
        return

    cg = capture_graphs[0]
    print(f"Using: {cg.get('name')}")

    # Build optimized index only for capture_graph's time range
    cg_start = cg["ts"]
    cg_end = cg_start + cg.get("dur", 0)
    cg_events = [
        e
        for e in capture_events
        if e.get("ph") == "X"
        and e.get("ts", 0) >= cg_start
        and e.get("ts", 0) + e.get("dur", 0) <= cg_end
    ]
    print(f"Events in capture_graph: {len(cg_events)}")
    idx = EventIndex(cg_events)

    # Get GPU kernels from first decode (within its duration)
    ds1_start = first_ds["ts"]
    ds1_end = ds1_start + first_ds.get("dur", 0)

    gpu_kernels = [
        e
        for e in run_events
        if e.get("cat") == "kernel" and ds1_start <= e["ts"] <= ds1_end
    ]
    gpu_kernels = sorted(gpu_kernels, key=lambda x: x["ts"])
    print(f"First decode (tid={first_ds.get('tid')}): {first_ds_name}")
    print(
        f"  Range: {ds1_start:.0f} ~ {ds1_end:.0f} (dur={first_ds.get('dur', 0):.0f})"
    )
    print(f"  GPU kernels: {len(gpu_kernels)}")

    # Use optimized index for children lookup
    direct_children = idx.get_direct_children(cg)
    kernel_children = [c for c in direct_children if idx.has_kernel_launch(c)]
    print(f"Direct children with kernels: {len(kernel_children)}")

    # Collect all modules with their kernel info
    all_modules = []  # list of (mod_name, kernel_count, start_gpu_idx)
    all_module_events = []
    gpu_idx = 0

    def _resolve_moe_child_name_decode(event: Dict[str, Any]) -> str:
        mod_name = event.get("name", "<unknown>")
        if "moe" not in mod_name.lower():
            return mod_name
        children = idx.get_direct_children(event)
        children_with_launch = [c for c in children if idx.has_kernel_launch(c)]
        if children_with_launch:
            return children_with_launch[0].get("name", mod_name)
        return mod_name

    for child in kernel_children:
        child_name = child.get("name", "")
        if should_filter(child_name):
            continue

        # Get sub-children (actual module names)
        sub_children = idx.get_direct_children(child)
        sub_kernel_children = [sc for sc in sub_children if idx.has_kernel_launch(sc)]

        # Determine modules to process
        modules = sub_kernel_children if sub_kernel_children else [child]

        for mod in modules:
            mod_name = _resolve_moe_child_name_decode(mod)
            kernel_count = idx.count_kernel_launches(mod)
            all_modules.append((mod_name, kernel_count, gpu_idx))
            all_module_events.append(mod)
            gpu_idx += kernel_count

    # Decode module sequence should start from the first norm module to avoid
    # wrapper/initialization nodes before the model block.
    first_norm_module_idx = next(
        (i for i, (name, _, _) in enumerate(all_modules) if is_strict_norm_name(name)),
        None,
    )
    if first_norm_module_idx is None:
        print("No norm module found in capture_graph modules.")
        return
    if first_norm_module_idx > 0:
        print(f"Dropped {first_norm_module_idx} leading modules before first norm.")
    all_modules = all_modules[first_norm_module_idx:]
    all_module_events = all_module_events[first_norm_module_idx:]

    # Anchor module->kernel alignment by first norm's correlated launch kernel:
    # 1) find first launch inside first norm and read its correlation/kernel
    # 2) find same kernel's first occurrence in run decode kernels
    # 3) rebuild module start indices from that anchor.
    capture_runtime_launches = [
        e
        for e in cg_events
        if e.get("cat") == "cuda_runtime" and is_kernel_launch(e.get("name", ""))
    ]
    capture_runtime_launches.sort(key=lambda x: x.get("ts", 0))
    capture_launch_ts = [e.get("ts", 0) for e in capture_runtime_launches]

    def _first_kernel_name_in_capture(mod_event: Dict[str, Any]) -> Optional[str]:
        m_start = mod_event.get("ts", 0)
        m_end = m_start + mod_event.get("dur", 0)
        left = bisect.bisect_left(capture_launch_ts, m_start)
        right = bisect.bisect_right(capture_launch_ts, m_end)
        for launch in capture_runtime_launches[left:right]:
            corr = (launch.get("args") or {}).get("correlation")
            if corr is None:
                continue
            kernel_name = (launch.get("args") or {}).get("kernel", "")
            if kernel_name:
                return str(kernel_name)
        return None

    anchor_kernel_name = _first_kernel_name_in_capture(all_module_events[0])
    if not anchor_kernel_name:
        raise RuntimeError(
            "Cannot resolve anchor kernel from first rmsnorm correlation in capture trace."
        )
    found = next(
        (
            i
            for i, k in enumerate(gpu_kernels)
            if k.get("name", "") == anchor_kernel_name
        ),
        None,
    )
    if found is None:
        raise RuntimeError(
            f"Anchor kernel '{anchor_kernel_name}' not found in run decode kernels."
        )
    anchor_gpu_idx = found
    print(
        f"Aligned from first norm kernel: {anchor_kernel_name} at gpu_idx={anchor_gpu_idx}"
    )

    rebuilt_modules = []
    running_gpu_idx = anchor_gpu_idx
    for name, count, _ in all_modules:
        rebuilt_modules.append((name, count, running_gpu_idx))
        running_gpu_idx += count
    all_modules = rebuilt_modules

    # Find norm positions (rmsnorm in name)
    all_norm_indices = [
        i for i, (name, _, _) in enumerate(all_modules) if is_strict_norm_name(name)
    ]

    def build_rows_for_module_range(start: int, end: int) -> List[List[Any]]:
        rows = []
        for mod_name, kernel_count, start_gpu_idx in all_modules[start:end]:
            rows.extend(
                process_module(mod_name, kernel_count, start_gpu_idx, gpu_kernels)
            )
        return rows

    _extract_layer_and_write(
        all_norm_indices,
        len(all_modules),
        target_layer,
        "Decode",
        "decode",
        build_rows_for_module_range,
        output_xlsx,
    )


# =============================================================================
# Main
# =============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Parse PyTorch profiler trace JSON to extract kernel information."
    )
    parser.add_argument("filepath", help="Path to trace JSON or JSON.GZ file")
    parser.add_argument(
        "--layer", type=int, default=3, help="Target layer index (default: 3)"
    )
    args = parser.parse_args()

    if args.layer < 0:
        print("--layer must be >= 0")
        sys.exit(1)

    filepath = args.filepath
    target_layer = args.layer

    print(f"Loading run trace: {filepath}")
    trace = load_trace(filepath)
    events = trace.get("traceEvents", [])
    print(f"Loaded run events: {len(events)}")

    capture_trace_path = find_capture_graph_trace_path(filepath)
    if capture_trace_path is None:
        print(
            "Warning: matching capture trace not found; decode analysis will fallback "
            "to current trace for capture_graph hierarchy."
        )
        capture_events = events
    else:
        print(f"Loading capture trace: {capture_trace_path}")
        capture_trace = load_trace(capture_trace_path)
        capture_events = capture_trace.get("traceEvents", [])
        print(f"Loaded capture events: {len(capture_events)}")
    print("")

    print("=" * 60)
    print("PREFILL ANALYSIS")
    print("=" * 60)
    parse_prefill(events, "prefill_breakdown.xlsx", target_layer=target_layer)

    print("\n" + "=" * 60)
    print("DECODE ANALYSIS")
    print("=" * 60)
    parse_decode(
        events,
        capture_events,
        "decode_breakdown.xlsx",
        target_layer=target_layer,
    )


if __name__ == "__main__":
    main()
